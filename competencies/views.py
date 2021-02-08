from django.shortcuts import render
from openpyxl import load_workbook
from myrir.utils import get_model_change_link

from .models import Competence, Category
from django.http import HttpResponseRedirect
from django.urls import path, reverse
from django.contrib import messages


# Create your views here.
def import_competencies(request):
    if request.method == 'POST' and 'competencies' in request.FILES:
        files = request.FILES['competencies']
        if files:
            book = load_workbook(files.file)
            sheet = book.active
            added_categories = []
            excluded_categories = []
            for i in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=i, column=1).value
                if cell_val:
                    try:
                        find_cat = Category.objects.get(name=str.strip(cell_val))
                        excluded_categories.append(find_cat)
                    except Category.DoesNotExist:
                        find_cat = None
                    if not find_cat:
                        new_category = Category.objects.create(name=cell_val)
                        # competencies = []
                        added_categories.append(new_category)
                        print(new_category)
                        competencies_list = sheet.cell(row=i, column=2).value.split(',')
                        for competence in competencies_list:
                            try:
                                finded_competence = Competence.objects.get(name=str.strip(competence))
                            except Competence.DoesNotExist:
                                finded_competence = None
                            if not finded_competence:
                                Competence.objects.create(name=str.strip(competence), category=new_category)
                                # competencies.append(new_competence)
                        # @TODO При большом количестве проектов в файле, будет происходить много запросов!
                        # Необходимо подумать возможно ли их привести к одному
            if added_categories:
                messages.success(
                    request,
                    f"Добавлены следующие категории: {', '.join(map(str, added_categories))}"
                )
            if excluded_categories:
                display_text = ", ".join([get_model_change_link(category) for category in excluded_categories])
                messages.warning(
                    request,
                    f"Следующие категории были добавлены ранее: {display_text}"
                )
    return HttpResponseRedirect('../')
