from django.contrib import admin
from openpyxl import load_workbook

from .models import Competence, Category
from django.http import HttpResponseRedirect
from django.urls import path

# Register your models here.
# admin.site.register(Competence)


class CompetenceInlineAdmin(admin.TabularInline):
    model = Competence
    extra = 1


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    change_list_template = 'admin/model_change_list.html'
    inlines = (CompetenceInlineAdmin,)

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import_competencies/', self.import_competencies, name='import_competencies'),
        ]
        return my_urls + urls

    def import_competencies(self, request):
        if request.method == 'POST' and 'competencies' in request.FILES:
            files = request.FILES['competencies']
            if files:
                book = load_workbook(files.file)
                sheet = book.active
                for i in range(2, sheet.max_row + 1):
                    cell_val = sheet.cell(row=i, column=1).value
                    if cell_val:
                        try:
                            find_cat = Category.objects.get(name=str.strip(cell_val))
                        except Category.DoesNotExist:
                            find_cat = None
                        if not find_cat:
                            new_category = Category.objects.create(name=cell_val)
                            competencies = []
                            competencies_list = sheet.cell(row=i, column=2).value.split(',')
                            for competence in competencies_list:
                                try:

                                    finded_competence = Competence.objects.get(name=str.strip(competence))
                                except Competence.DoesNotExist:
                                    finded_competence = None
                                if not finded_competence:
                                    new_competence = Competence.objects.create(name=str.strip(competence), category=new_category)
                                    competencies.append(new_competence)
                            # @TODO При большом количестве проектов в файле, будет происходить много запросов!
                            # Необходимо подумать возможно ли их привести к одному
        return HttpResponseRedirect('../')
