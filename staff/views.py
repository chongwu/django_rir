import datetime

from django.shortcuts import render, get_object_or_404, HttpResponse, redirect
from django.utils.encoding import smart_str
from django.http import HttpResponseRedirect
from django.core import serializers
from django.contrib.auth.models import User
from .models import Person, Department, Position, Staff, EMPLOYMENT_FORMS_DICT, WORK_STATUSES_DICT
from competencies.models import Category
from questionnaires.models import Questionnaire, QuestionnaireRow, QuestionnaireRowHistory
from itertools import groupby
from operator import attrgetter
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from tempfile import NamedTemporaryFile
from django.contrib import messages
from myrir.utils import group_required
from dateutil.relativedelta import relativedelta
import os
import xlwt


# Create your views here.
def person_list(request, department_id):
    department = Department.objects.get(pk=department_id)
    # persons = Person.objects.all()
    persons = department.workers.filter(person_staff__date_stop__isnull=True)
    return render(request, 'staff/person/list.html', {'department': department, 'persons': persons})


def departments_list(request):
    departments = Department.objects.all()
    return render(request, 'staff/department/list.html', {'departments': departments})


def person_detail(request, person_id):
    person = get_object_or_404(Person.objects.select_related('map'), tab_number=person_id)
    questionnaire = person.questionnaires.prefetch_related('rows', 'rows__competence', 'rows__competence__category').first()
    history = QuestionnaireRowHistory.objects.filter(questionnaire_row__questionnaire=questionnaire)\
        .prefetch_related('questionnaire_row__competence').all()
    rows = {}
    if questionnaire:
        rows = {k: list(v) for k, v in groupby(questionnaire.rows.all(), attrgetter('competence.category.name'))}
    return render(request, 'staff/person/detail.html', {'person': person, 'rows': rows, 'history': history})


@group_required('HR')
def create_questionnaire(request, person_id):
    person = get_object_or_404(Person, tab_number=person_id)
    categories = person.position.first().competence_category.all()
    return render(request, 'staff/person/questionnaire_form.html', {'person': person, 'categories': categories})


@group_required('HR')
def upload_questionnaire(request, person_id):
    person = get_object_or_404(Person, tab_number=person_id)
    if request.method == 'POST':
        files = request.FILES['questionnaire']
        if files:
            book = load_workbook(files.file)
            sheet = book.active
            questionnaire = Questionnaire(person=person)
            questionnaire.save()
            rows = []
            for i in range(1, sheet.max_row + 1):
                rows.append(QuestionnaireRow(
                    questionnaire=questionnaire,
                    competence_id=sheet.cell(row=i, column=1).value,
                    competence_val=sheet.cell(row=i, column=3).value if sheet.cell(row=i, column=3).value else 0,
                ))
            QuestionnaireRow.objects.bulk_create(rows)
            messages.success(request, 'Анкета сотрудника успешно загружена!')
    return redirect(person)


@group_required('HR')
def download_questionnaire(request, person_id):
    person = get_object_or_404(Person, tab_number=person_id)
    categories = person.position.first().competence_category.all()
    with NamedTemporaryFile(mode='w', delete=False) as f:
        book = Workbook()
        sheet = book.active
        i = 1
        for category in categories:
            for competence in category.competencies.all():
                sheet.cell(row=i, column=1).value = competence.id
                sheet.cell(row=i, column=2).value = competence.name
                i += 1
        book.save(f.name)
        book.close()
        file_name = os.path.basename(f.name)
        response = HttpResponse(content=save_virtual_workbook(book), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(file_name + '.xlsx')
    return response


@group_required('HR')
def upload_department_list(request):
    if request.method == 'POST' and 'departments' in request.FILES:
        files = request.FILES['departments']
        if files:
            book = load_workbook(files.file)
            sheet = book.active
            rows = []
            for i in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=i, column=1).value
                if cell_val:
                    rows.append(Department(
                        name=cell_val,
                    ))
            Department.objects.bulk_create(rows)
            messages.success(request, 'Отделы успешно загружены!')
    return redirect('staff:departments_list')


@group_required('HR')
def upload_positions_list(request):
    if request.method == 'POST' and 'positions' in request.FILES:
        files = request.FILES['positions']
        book = load_workbook(files.file)
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            cell_val = sheet.cell(row=i, column=1).value
            if cell_val:
                new_position = Position.objects.create(
                    name=cell_val
                )
                competence_categories = []
                categories_list = sheet.cell(row=i, column=2).value.split(',')
                for category in categories_list:
                    finded_category = Category.objects.get(name=str.strip(category))
                    if finded_category:
                        competence_categories.append(finded_category)
                new_position.competence_category.add(*competence_categories)
        messages.success(request, 'Список должностей успешно загружен!')
    return redirect('staff:departments_list')


@group_required('HR')
def upload_persons_list(request):
    if request.method == 'POST' and 'persons' in request.FILES:
        files = request.FILES['persons']
        # TODO Добавить обработку ошибок, если загружается не xlsx файл
        book = load_workbook(files.file)
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            cell_val = sheet.cell(row=i, column=1).value
            if cell_val:
                try:
                    exist_person = Person.objects.get(pk=cell_val)
                except Person.DoesNotExist:
                    exist_person = None
                if not exist_person:
                    # TODO Преобразовать в функцию!
                    # try:
                    #     user = User.objects.get(username=sheet.cell(row=i, column=2).value)
                    # except User.DoesNotExist:
                    #     user = None
                    # try:
                    #     mentor = Person.objects.get(fio=sheet.cell(row=i, column=10).value)
                    # except Person.DoesNotExist:
                    #     mentor = None

                    start_date = sheet.cell(row=i, column=7).value
                    if datetime.datetime.now() - relativedelta(months=3) > start_date:
                        status = 2
                    else:
                        status = 1
                    # self.employee.person_staff.get().date_start + relativedelta(months=3)

                    new_person = Person.objects.create(
                        tab_number=cell_val,
                        fio=sheet.cell(row=i, column=2).value,
                        # user=user,
                        education=sheet.cell(row=i, column=3).value,
                        institution=sheet.cell(row=i, column=4).value,
                        # experience=sheet.cell(row=i, column=6).value,
                        # extra_skill=sheet.cell(row=i, column=7).value,
                        # employment_form=list(EMPLOYMENT_FORMS_DICT.keys())[list(EMPLOYMENT_FORMS_DICT.values()).index(sheet.cell(row=i, column=8).value)],
                        # status=list(WORK_STATUSES_DICT.keys())[list(WORK_STATUSES_DICT.values()).index(sheet.cell(row=i, column=9).value)],
                        status=status,
                        # mentor=mentor,
                    )

                    try:
                        position = Position.objects.get(name=sheet.cell(row=i, column=5).value)
                    except Position.DoesNotExist:
                        position = None
                    try:
                        department = Department.objects.get(name=sheet.cell(row=i, column=6).value)
                    except Department.DoesNotExist:
                        department = None

                    if position and department:
                        Staff.objects.create(
                            person=new_person,
                            position=position,
                            department=department,
                            date_start=start_date.strftime('%Y-%m-%d')
                        )

        messages.success(request, 'Список сотрудников успешно загружен!')
    # return redirect('staff:person_list', department_id)
    return HttpResponseRedirect('../')


def json_users(request):
    data = serializers.serialize("json", User.objects.all(), fields=('username',))
    return HttpResponse(data, content_type='application/json')


def json_persons(request):
    data = serializers.serialize("json", Person.objects.all(), fields=('fio',))
    return HttpResponse(data, content_type='application/json')


def json_departments(request):
    data = serializers.serialize("json", Department.objects.all(), fields=('name',))
    return HttpResponse(data, content_type='application/json')


def json_positions(request):
    data = serializers.serialize("json", Position.objects.all(), fields=('name',))
    return HttpResponse(data, content_type='application/json')

# @TODO 13.12.2021 - остановился здесь

def export_employees_xlsx(request, department_id=None):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="employees.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Сотрудники')


