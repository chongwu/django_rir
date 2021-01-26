from django.shortcuts import render, redirect
from openpyxl import load_workbook

from .models import Product, Project
from staff.models import Person


# Create your views here.
def products_list(request):
    products = Product.objects.all()
    return render(request, 'products/products/list.html', {'products': products})


def view_product(request, product_id):
    product = Product.objects.filter(pk=product_id).prefetch_related('projects').get()
    return render(request, 'products/products/detail.html', {'product': product})


def view_project(request, product_id, project_id):
    product = Product.objects.get(pk=product_id)
    project = Project.objects.filter(pk=project_id).prefetch_related('persons', 'persons__position').get()
    return render(request, 'products/projects/detail.html', {'project': project, 'product': product})


def upload_products_list(request):
    if request.method == 'POST' and 'products' in request.FILES:
        files = request.FILES['products']
        if files:
            book = load_workbook(files.file)
            sheet = book.active
            rows = []
            for i in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=i, column=1).value
                if cell_val:
                    rows.append(Product(
                        name=cell_val,
                        region=sheet.cell(row=i, column=2).value,
                        description=sheet.cell(row=i, column=3).value,
                        additional_info=sheet.cell(row=i, column=4).value,
                    ))
            Product.objects.bulk_create(rows)
    return redirect('products:all_products')


def upload_projects_list(request, product_id):
    if request.method == 'POST' and 'projects' in request.FILES:
        files = request.FILES['projects']
        if files:
            book = load_workbook(files.file)
            sheet = book.active
            # rows = []
            for i in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=i, column=1).value
                if cell_val:
                    persons = []
                    persons_list = sheet.cell(row=i, column=5).value.split(',')
                    for person in persons_list:
                        try:
                            finded_person = Person.objects.get(fio=str.strip(person))
                        except Person.DoesNotExist:
                            finded_person = None
                        if finded_person:
                            persons.append(finded_person)
                    # @TODO При большом количестве проектов в файле, будет происходить много запросов!
                    # Необходимо подумать возможно ли их привести к одному
                    new_project = Project.objects.create(
                        name=cell_val,
                        city=sheet.cell(row=i, column=2).value,
                        description=sheet.cell(row=i, column=3).value,
                        additional_info=sheet.cell(row=i, column=4).value,
                        product_id=product_id
                    )
                    new_project.persons.add(*persons)
    return redirect('products:view_product', product_id)
